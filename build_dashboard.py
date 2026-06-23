"""
Courage Team Performance Dashboard — Weekly Build Script
=========================================================

Reads source data and generates the updated HTML dashboard.

EXPECTED FOLDER STRUCTURE:

    Team Performance Dashboard/
    ├── build_dashboard.py                          ← this script
    ├── template_dashboard.html                     ← the HTML template (do not delete)
    ├── Courage_Team_Performance_Dashboard.html     ← OUTPUT (overwritten each run)
    └── Data Source/
        └── NCC_TEAM_Match_Data.xlsx                ← update weekly

WEEKLY UPDATE WORKFLOW:
  1. Add the new match's row(s) to NCC_TEAM_Match_Data.xlsx
     (one NCC row [Opponent Data blank] + one Opp row [Opponent Data = TRUE])
  2. Run this script (in Cowork: "Run build_dashboard.py")
  3. Open Courage_Team_Performance_Dashboard.html

DATA SOURCES:
  Spreadsheet (auto)  → All metrics + match metadata + match result
                        (computed from Goals column) + Tackles Won +
                        Opp Possession %
  MANUAL_OVERRIDES    → Post-Shot xG fallback (auto-read from spreadsheet
                        "Post Shot xG" column), plus Match Insights fields:
                        Shots on Target (ncc_sot/opp_sot), Press Regains
                        (ncc_press_regains/opp_press_regains), and narrative
                        text (top_shooters, top_kp, note).
"""

import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).resolve().parent
DATA_DIR = SCRIPT_DIR / "Data Source"
SPREADSHEET = DATA_DIR / "NWSL Match Data - Team Level.xlsx"
SPREADSHEET_SHEET = "NCC Data"
TEMPLATE = SCRIPT_DIR / "template_dashboard.html"
OUTPUT = SCRIPT_DIR / "Courage_Team_Performance_Dashboard.html"

# ─────────────────────────────────────────────────────────────────────────────
# MANUAL OVERRIDES
# ─────────────────────────────────────────────────────────────────────────────
# Fields with no automatic spreadsheet source. Indexed by Game Number (M1, M2, …).
# When you add a new match, append a new key block (e.g. "M11": {...}).
# Use None for any value not yet available — dashboard will show "pending".
#
# Auto-sourced from spreadsheet (no manual entry needed):
#   psxg_ncc / psxg_opp  — "Post Shot xG" column (NCC + OPP rows)
#   All other DATA metrics (xG, Shots, Possession %, Passing%, Pressures, etc.)
#   Match result (computed from Goals column, NCC vs OPP rows)
#   Tackles Won (from Tackles column, NCC row)
#
# Requires manual entry per match (not in spreadsheet):
#   ncc_sot / opp_sot          — Shots on Target
#   ncc_press_regains / opp_press_regains — Press Regains
#   top_shooters / top_kp / note — narrative text for Match Insights cards
#
# If a Game Number is missing from this dict, all overridable metrics for that
# match default to None and the dashboard displays them as pending.

MANUAL_OVERRIDES = {
    # Game Number → { psxg_ncc, psxg_opp, ncc_sot, opp_sot,
    #                 ncc_press_regains, opp_press_regains,
    #                 top_shooters, top_kp, note }
    # NOTE: psxg_ncc/psxg_opp are auto-read from the spreadsheet
    # "Post Shot xG" column. Values here serve only as fallback.
    "M1": {
        "psxg_ncc": None, "psxg_opp": None,
        "ncc_sot": 5,  "opp_sot": 5,
        "ncc_press_regains": 56, "opp_press_regains": 53,
        "top_shooters": "Ashley Nicole Sanchez (5 shots, 0.25xG), Hannah Betfort (2, 0.64xG)",
        "top_kp": "Shinomi Koyama (3 KP, 0.11xG), Ryan Williams (2 KP, 0.17xG)",
        "note": "Won despite being xG-negative. Efficiency from open play key — 1.22xG from open play sources.",
    },
    "M2": {
        "psxg_ncc": None, "psxg_opp": None,
        "ncc_sot": 3,  "opp_sot": 1,
        "ncc_press_regains": 67, "opp_press_regains": 65,
        "top_shooters": None, "top_kp": None,
        "note": "We dominated xG and territory but failed to convert. Their GK held firm.",
    },
    "M3": {
        "psxg_ncc": None, "psxg_opp": None,
        "ncc_sot": 2,  "opp_sot": 6,
        "ncc_press_regains": 74, "opp_press_regains": 69,
        "top_shooters": None, "top_kp": None,
        "note": "Lost despite higher xG. Bay FC converted 3 from 6 SOT — clinical finishing overturned the chance-quality gap.",
    },
    "M4": {
        "psxg_ncc": None, "psxg_opp": None,
        "ncc_sot": 6,  "opp_sot": 6,
        "ncc_press_regains": 54, "opp_press_regains": 57,
        "top_shooters": None, "top_kp": None,
        "note": "Allowed 6 SOT despite low opposition xG.",
    },
    "M5": {
        "psxg_ncc": None, "psxg_opp": None,
        "ncc_sot": 8,  "opp_sot": 2,
        "ncc_press_regains": 51, "opp_press_regains": 58,
        "top_shooters": "Manaka Matsukubo (8 shots, 0.76xG)", "top_kp": "Avery Patterson (4 KP, 0.43xG)",
        "note": "Won on fine margins. Houston dominated set-piece xG (0.65) but we nullified it. Matsukubo led our attacking threat.",
    },
    "M6": {
        "psxg_ncc": None, "psxg_opp": None,
        "ncc_sot": 3,  "opp_sot": 2,
        "ncc_press_regains": 63, "opp_press_regains": 78,
        "top_shooters": None, "top_kp": None,
        "note": "We scored 2 from just 0.78xG — season-high over-performance. Boston had 1.85xG but we held to a draw.",
    },
    "M7": {
        "psxg_ncc": None, "psxg_opp": None,
        "ncc_sot": 3,  "opp_sot": 2,
        "ncc_press_regains": 37, "opp_press_regains": 33,
        "top_shooters": None, "top_kp": None,
        "note": "Season low in pressures (93). KC had season-high opp xG (2.18). Low press intensity linked to defensive exposure.",
    },
    "M8": {
        "psxg_ncc": None, "psxg_opp": None,
        "ncc_sot": 1,  "opp_sot": 2,
        "ncc_press_regains": 37, "opp_press_regains": 29,
        "top_shooters": None, "top_kp": None,
        "note": "Lost despite leading on xG (0.78 vs 0.45) and possession (57%). Orlando converted their only quality chance — we were clinical in neither box.",
    },
    "M9": {
        "psxg_ncc": None, "psxg_opp": None,
        "ncc_sot": None, "opp_sot": None,
        "ncc_press_regains": None, "opp_press_regains": None,
        "top_shooters": None, "top_kp": None,
        "note": None,
    },
    "M10": {
        "psxg_ncc": None, "psxg_opp": None,
        "ncc_sot": 4,  "opp_sot": 2,
        "ncc_press_regains": 44, "opp_press_regains": 48,
        "top_shooters": "Ashley Nicole Sanchez (4 shots, 0.45xG), Shinomi Koyama (2, 0.09xG), Felicitas Rauch (2, 0.11xG)",
        "top_kp": "Ryan Williams (2 KP, 0.67xG), Allyson Schlegel (2 KP, 0.11xG)",
        "note": "W 2-1 away at Louisville. Won despite lower xG (1.23 vs 1.29) — clinical finishing from ANS. Louisville's corner volume generated 0.60 SP xG but NCC held firm defensively.",
    },
    "M11": {
        "psxg_ncc": None, "psxg_opp": None,
        "ncc_sot": None, "opp_sot": None,
        "ncc_press_regains": None, "opp_press_regains": None,
        "top_shooters": None, "top_kp": None,
        "note": None,
    },
}


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def warn(msg):
    print(f"  ⚠️  {msg}")

def info(msg):
    print(f"  · {msg}")


def safe_round(v, dec):
    if v is None:
        return None
    return round(float(v), dec)


def safe_div(num, den):
    if num is None or den is None or den == 0:
        return None
    return num / den


# ─────────────────────────────────────────────────────────────────────────────
# 1) READ SPREADSHEET
# ─────────────────────────────────────────────────────────────────────────────

def read_spreadsheet():
    """
    Returns (matches, sheet_data) where:
      matches    = ordered list of match metadata dicts (includes 'result')
      sheet_data = dict mapping (game_number, side) → {column_name: value}
                   side ∈ {"NCC", "OPP"}
    """
    print(f"\n[1/3] Reading spreadsheet: {SPREADSHEET.name}")
    if not SPREADSHEET.exists():
        sys.exit(f"ERROR: Spreadsheet not found at {SPREADSHEET}")

    wb = openpyxl.load_workbook(SPREADSHEET, data_only=True)
    ws = wb[SPREADSHEET_SHEET]

    # Build header → column index map
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None:
            headers[str(v).strip()] = c

    # Required headers
    for required in ("Date", "Game Week", "Game Number", "Match", "Team"):
        if required not in headers:
            sys.exit(f"ERROR: Required column '{required}' missing from spreadsheet")

    matches = []         # ordered list of {game_number, game_week, date, opp, loc, ncc_home, result}
    sheet_data = {}      # (game_number, "NCC"|"OPP") → row dict

    seen_gn = set()
    for r in range(2, ws.max_row + 1):
        date_v = ws.cell(row=r, column=headers["Date"]).value
        if date_v is None:
            continue
        # Date may be a datetime or a string like "3/15/26"
        if isinstance(date_v, str):
            parsed = None
            for fmt in ("%m/%d/%y", "%m/%d/%Y", "%Y-%m-%d"):
                try:
                    parsed = datetime.strptime(date_v.strip(), fmt)
                    break
                except ValueError:
                    continue
            if parsed is None:
                sys.exit(f"ERROR: Could not parse Date '{date_v}' on row {r}")
            date_v = parsed

        gw = ws.cell(row=r, column=headers["Game Week"]).value
        gn = ws.cell(row=r, column=headers["Game Number"]).value
        match = ws.cell(row=r, column=headers["Match"]).value or ""
        # Side is determined by the Team column: NCC rows have Team == North Carolina Courage
        team_v = ws.cell(row=r, column=headers["Team"]).value or ""
        is_opp_row = "north carolina courage" not in str(team_v).lower()

        # Read the entire row as a dict
        row_dict = {}
        for hdr, col in headers.items():
            row_dict[hdr] = ws.cell(row=r, column=col).value

        side = "OPP" if is_opp_row else "NCC"
        sheet_data[(gn, side)] = row_dict

        if not is_opp_row and gn not in seen_gn:
            seen_gn.add(gn)
            # Parse opponent name + home/away from "Match" cell
            m = re.match(r"\s*(.+?)\s*vs\.?\s*(.+?)\s*$", str(match))
            if m:
                t1, t2 = m.group(1).strip(), m.group(2).strip()
                if "north carolina courage" in t1.lower():
                    opp_name, ncc_home = t2, True
                else:
                    opp_name, ncc_home = t1, False
            else:
                opp_name, ncc_home = "Unknown", True

            matches.append({
                "game_number": gn,
                "game_week": gw,
                "date": date_v,
                "opp": opp_name,
                "loc": "Home" if ncc_home else "Away",
                "ncc_home": ncc_home,
                "match_string": str(match),
                "result": None,   # filled in after both rows are loaded
            })

    matches.sort(key=lambda m: m["game_number"])

    # Compute match result from Goals column (NCC row vs OPP row)
    for m in matches:
        gn = m["game_number"]
        ncc_goals = sheet_data.get((gn, "NCC"), {}).get("Goals")
        opp_goals = sheet_data.get((gn, "OPP"), {}).get("Goals")
        if ncc_goals is not None and opp_goals is not None:
            ng, og = int(ncc_goals), int(opp_goals)
            if ng > og:
                m["result"] = f"W {ng}-{og}"
            elif ng < og:
                m["result"] = f"L {ng}-{og}"
            else:
                m["result"] = f"D {ng}-{og}"
        else:
            warn(f"M{gn}: Goals column missing or blank — result will show as '—'")
            m["result"] = None

    info(f"Loaded {len(matches)} matches (Game Numbers: {[m['game_number'] for m in matches]})")
    for m in matches:
        info(f"  M{m['game_number']} ({m['opp']:<22}) {m['result'] or '?'}")
    return matches, sheet_data


# ─────────────────────────────────────────────────────────────────────────────
# 2) BUILD DATA STRUCTURES
# ─────────────────────────────────────────────────────────────────────────────

def get_opp_possession_pct(matches, sheet_data):
    """Opp Possession % per match. Resolution order:
       1. NCC row's "Opp Possession %" / "Opposition Possession %" column (single-value style)
       2. OPP row's "Possession %" column (per-team style — preferred new layout)
       3. Default 50% (with warning)
       Values stored as 0–1; sheet may store as 42 (percent) or 0.42 (decimal)."""
    out = []
    for m in matches:
        gn = m["game_number"]
        ncc_row = sheet_data.get((gn, "NCC"), {})
        opp_row = sheet_data.get((gn, "OPP"), {})
        v = None

        # 1) Single-value-style headers on the NCC row (e.g. "Opp Possession %")
        for hdr in ("Opp Possession %", "Opposition Possession %", "Opp. Possession %",
                    "Opp Possession", "Opposition Possession"):
            if hdr in ncc_row and ncc_row[hdr] is not None:
                v = ncc_row[hdr]
                break

        # 2) Per-team-style headers on the OPP row (e.g. just "Possession %")
        if v is None:
            for hdr in ("Possession %", "Possession%", "Possession"):
                if hdr in opp_row and opp_row[hdr] is not None:
                    v = opp_row[hdr]
                    break

        # 3) Default
        if v is None:
            warn(f"M{gn}: no Opp Possession % found in spreadsheet — defaulting to 0.50")
            v = 0.50

        # Normalize to 0–1
        v = float(v)
        if v > 1:
            v = v / 100.0
        out.append(v)
    return out


def pa_adjust(raw_value, opp_poss_pct):
    """PA formula: raw ÷ opp_possession_pct × 0.5"""
    if raw_value is None or opp_poss_pct is None or opp_poss_pct == 0:
        return None
    return raw_value / opp_poss_pct * 0.5


def build_data_block(matches, sheet_data):
    """Produce the DATA dict in dashboard format."""
    print(f"\n[2/3] Building DATA block")

    n_matches = len(matches)
    opp_poss = get_opp_possession_pct(matches, sheet_data)

    def vals_from(side, header, transform=None):
        """Extract values for a metric across all matches from a given side."""
        out = []
        for m in matches:
            row = sheet_data.get((m["game_number"], side), {})
            v = row.get(header)
            if transform is not None and v is not None:
                v = transform(v)
            out.append(v)
        return out

    def manual_vals(field):
        """Pull a manually-overridden field across all matches.
        For Post-Shot xG fields (psxg_ncc, psxg_opp), the spreadsheet
        'Post Shot xG' column is checked first (NCC row for psxg_ncc,
        OPP row for psxg_opp). MANUAL_OVERRIDES values are used only as
        a fallback when the spreadsheet column is absent or blank."""
        PSXG_SHEET_COL = "Post Shot xG"
        PSXG_FIELD_TO_SIDE = {"psxg_ncc": "NCC", "psxg_opp": "OPP"}
        out = []
        for m in matches:
            key = f"M{m['game_number']}"
            # Prefer spreadsheet column for Post-Shot xG
            if field in PSXG_FIELD_TO_SIDE:
                side = PSXG_FIELD_TO_SIDE[field]
                row = sheet_data.get((m["game_number"], side), {})
                sheet_val = row.get(PSXG_SHEET_COL)
                if sheet_val is not None:
                    out.append(sheet_val)
                    continue
            # Fall back to MANUAL_OVERRIDES
            entry = MANUAL_OVERRIDES.get(key, {})
            out.append(entry.get(field))
        return out

    # ─── SOP ────────────────────────────────────────────────────────────────
    z2_att = vals_from("NCC", "Z2 Attempts")
    z2_comp = vals_from("NCC", "Z2 Completions")
    z2_pct = [safe_div(c, a) for c, a in zip(z2_comp, z2_att)]
    z3_att = vals_from("NCC", "Z3 Attempts")
    z3_comp = vals_from("NCC", "Z3 Completions")
    z3_pct = [safe_div(c, a) for c, a in zip(z3_comp, z3_att)]
    pz_att = vals_from("NCC", "PZ Attempts")
    pz_comp = vals_from("NCC", "PZ Completions")
    pz_pct = [safe_div(c, a) for c, a in zip(pz_comp, pz_att)]
    sz_att = vals_from("NCC", "SZ Attempts")
    sz_comp = vals_from("NCC", "SZ Completions")
    sz_pct = [safe_div(c, a) for c, a in zip(sz_comp, sz_att)]

    # ─── Packing ────────────────────────────────────────────────────────────
    byp_opp = vals_from("NCC", "Bypassed Opponents")
    byp_def = vals_from("NCC", "Bypassed Defenders")
    bw_opp = vals_from("NCC", "Ball Win Removed Opponents")
    bw_def_raw = vals_from("NCC", "Ball Win Removed Opponents Defenders")
    bw_def = [pa_adjust(v, p) for v, p in zip(bw_def_raw, opp_poss)]
    crit_loss = vals_from("NCC", "Critical Ball Loss Number")
    bl_team = vals_from("NCC", "Ball Loss Removed Teammates")
    deep_prog = vals_from("NCC", "Deep Progressions")
    lb_comp_f3 = vals_from("NCC", "Line Breaking Passes Completed in Final Third")
    lb_pct_f3 = vals_from("NCC", "Line Breaking Passes Completed in Final Third%")
    lb_obv_f3 = vals_from("NCC", "Line Breaking Passes On Ball Value in Final Third")

    # ─── Box Domination ─────────────────────────────────────────────────────
    bea_cp = vals_from("NCC", "Box Entries Attempted (Cross&Pass)")
    bea_carry = vals_from("NCC", "Box Entries Attempted (Carry)")
    sbe = vals_from("NCC", "Successful Box Entries")
    tib = vals_from("NCC", "Touches in box")
    pitb = vals_from("NCC", "Passes Into Box")
    spitb = vals_from("NCC", "Successful Passes Into Box")
    sbc = vals_from("NCC", "Successful Box Crosses")
    xg = vals_from("NCC", "xG")
    spxg = vals_from("NCC", "Set Piece xG")
    shots = vals_from("NCC", "Shots")
    xgps = vals_from("NCC", "xG/Shot")
    psxg = manual_vals("psxg_ncc")
    qc = vals_from("NCC", "Quality Chances")
    bc_v = vals_from("NCC", "Big Chances")

    # ─── Box Resilience ─────────────────────────────────────────────────────
    opp_bea_cp = vals_from("OPP", "Box Entries Attempted (Cross&Pass)")
    opp_bea_carry = vals_from("OPP", "Box Entries Attempted (Carry)")
    opp_sbe = vals_from("OPP", "Successful Box Entries")
    opp_tib = vals_from("OPP", "Touches in box")
    opp_pitb = vals_from("OPP", "Passes Into Box")
    opp_spitb = vals_from("OPP", "Successful Passes Into Box")
    opp_sbc = vals_from("OPP", "Successful Box Crosses")
    opp_xg = vals_from("OPP", "xG")
    opp_spxg = vals_from("OPP", "Set Piece xG")
    opp_shots = vals_from("OPP", "Shots")
    opp_xgps = vals_from("OPP", "xG/Shot")
    opp_psxg = manual_vals("psxg_opp")
    opp_qc = vals_from("OPP", "Quality Chances")
    opp_bc_v = vals_from("OPP", "Big Chances")

    # ─── Defensive Intensity ────────────────────────────────────────────────
    pressures_raw = vals_from("NCC", "Pressures")
    cp_raw = vals_from("NCC", "Counterpressures")
    agg_raw = vals_from("NCC", "Aggressive Actions")
    # Tackles Won — read from the spreadsheet's 'Tackles' column (NCC row).
    tackles_won = vals_from("NCC", "Tackles")
    interceptions_raw = vals_from("NCC", "Interceptions")
    tack_int_raw = [
        (t or 0) + (i or 0) if (t is not None or i is not None) else None
        for t, i in zip(tackles_won, interceptions_raw)
    ]
    pressures = [pa_adjust(v, p) for v, p in zip(pressures_raw, opp_poss)]
    cp = [pa_adjust(v, p) for v, p in zip(cp_raw, opp_poss)]
    agg_actions = [pa_adjust(v, p) for v, p in zip(agg_raw, opp_poss)]
    tack_int = [pa_adjust(v, p) for v, p in zip(tack_int_raw, opp_poss)]

    press_oh_raw = vals_from("NCC", "Pressures in Opposing Half")
    press_oh = [pa_adjust(v, p) for v, p in zip(press_oh_raw, opp_poss)]
    press_oh_pct = vals_from("NCC", "Pressures in Opposing Half%")
    cp_oh_raw = vals_from("NCC", "Counterpressures in Opposing Half")
    cp_oh = [pa_adjust(v, p) for v, p in zip(cp_oh_raw, opp_poss)]

    opp_dp = vals_from("OPP", "Deep Progressions")
    opp_lb_comp_f3 = vals_from("OPP", "Line Breaking Passes Completed in Final Third")
    opp_lb_pct = vals_from("OPP", "Line Breaking Passes Completed in Final Third%")

    # ─── On-Ball Value (OBV) ────────────────────────────────────────────────
    obv_total      = vals_from("NCC", "OBV")
    obv_pass       = vals_from("NCC", "Pass OBV")
    obv_shot       = vals_from("NCC", "Shot OBV")
    obv_dribble    = vals_from("NCC", "Dribble & Carry OBV")
    obv_def        = vals_from("NCC", "Defensive Action OBV")

    # Opposition OBV lives on the NCC row under "Opposition ..." column headers
    opp_obv_total  = vals_from("NCC", "Opposition OBV")
    opp_obv_pass   = vals_from("NCC", "Opposition Pass OBV")
    opp_obv_shot   = vals_from("NCC", "Opposition Shot OBV")
    opp_obv_dribble= vals_from("NCC", "Opposition Dribble & Carry OBV")
    opp_obv_def    = vals_from("NCC", "Opposition Defensive Action OBV")

    # Round to dashboard's decimal preferences
    def r(arr, dec):
        return [safe_round(v, dec) for v in arr]

    DATA = {
        "sop": {
            "progression": [
                {"id": "Z2_att",  "label": "Zone 2 Attempts",     "vals": z2_att,  "higherBetter": True},
                {"id": "Z2_comp", "label": "Zone 2 Completions",  "vals": z2_comp, "higherBetter": True},
                {"id": "Z2_pct",  "label": "Zone 2 Completion %", "vals": r(z2_pct, 4), "higherBetter": True, "pct": True, "dec": 3},
                {"id": "Z3_att",  "label": "Zone 3 Attempts",     "vals": z3_att,  "higherBetter": True},
                {"id": "Z3_comp", "label": "Zone 3 Completions",  "vals": z3_comp, "higherBetter": True},
                {"id": "Z3_pct",  "label": "Zone 3 Completion %", "vals": r(z3_pct, 4), "higherBetter": True, "pct": True, "dec": 3},
            ],
            "finishing": [
                {"id": "PZ_att",  "label": "Pass Zone Attempts",     "vals": pz_att,  "higherBetter": True},
                {"id": "PZ_comp", "label": "Pass Zone Completions",  "vals": pz_comp, "higherBetter": True},
                {"id": "PZ_pct",  "label": "Pass Zone Completion %", "vals": r(pz_pct, 4), "higherBetter": True, "pct": True, "dec": 3},
                {"id": "SZ_att",  "label": "Shoot Zone Attempts",     "vals": sz_att,  "higherBetter": True},
                {"id": "SZ_comp", "label": "Shoot Zone Completions",  "vals": sz_comp, "higherBetter": True},
                {"id": "SZ_pct",  "label": "Shoot Zone Completion %", "vals": r(sz_pct, 4), "higherBetter": True, "pct": True, "dec": 3},
            ],
        },
        "packing": {
            "volume": [
                {"id": "byp_opp",   "label": "Bypassed Opponents",          "vals": byp_opp,   "higherBetter": True},
                {"id": "byp_def",   "label": "Bypassed Defenders",          "vals": byp_def,   "higherBetter": True},
                {"id": "bw_opp",    "label": "Ball Win Removed Opponents",  "vals": bw_opp,    "higherBetter": True},
                {"id": "bw_def",    "label": "Ball Win Removed Defenders (PA)", "vals": r(bw_def, 1), "higherBetter": True, "dec": 1},
                {"id": "crit_loss", "label": "Critical Ball Loss",          "vals": crit_loss, "higherBetter": False, "context": "neg"},
                {"id": "bl_team",   "label": "Ball Loss Removed Teammates", "vals": bl_team,   "higherBetter": False, "context": "neg"},
                {"id": "deep_prog", "label": "Final 3rd Entries",           "vals": deep_prog, "higherBetter": True},
            ],
            "linebreak": [
                {"id": "lb_comp_f3",  "label": "LB Passes Completed in Final Third",   "vals": lb_comp_f3,            "higherBetter": True},
                {"id": "lb_pct_f3",   "label": "LB Completion % (Final Third)",        "vals": r(lb_pct_f3, 4),       "higherBetter": True, "pct": True, "dec": 3},
                {"id": "lb_obv_f3",   "label": "LB Passes OBV (Final Third)",          "vals": r(lb_obv_f3, 3),       "higherBetter": True, "dec": 3},
            ],
        },
        "boxdom": {
            "penetration": [
                {"id": "bea_cp",    "label": "Box Entries Attempted (Cross&Pass)",  "vals": bea_cp,    "higherBetter": True},
                {"id": "bea_carry", "label": "Box Entries Attempted (Carry)",       "vals": bea_carry, "higherBetter": True},
                {"id": "sbe",       "label": "Successful Box Entries",              "vals": sbe,       "higherBetter": True},
                {"id": "tib",       "label": "Touches in Box",                      "vals": tib,       "higherBetter": True},
            ],
            "finishing": [
                {"id": "xg",    "label": "Total xG",       "vals": r(xg, 3),    "higherBetter": True, "dec": 3},
                {"id": "spxg",  "label": "Set Piece xG",   "vals": r(spxg, 3),  "higherBetter": True, "dec": 3},
                {"id": "shots", "label": "Shots",          "vals": shots,       "higherBetter": True},
                {"id": "xgps",  "label": "xG/Shot",        "vals": r(xgps, 3),  "higherBetter": True, "dec": 3},
                {"id": "psxg",  "label": "Post-Shot xG",   "vals": r(psxg, 2),  "higherBetter": True, "dec": 2},
                {"id": "qc",    "label": "Quality Chances","vals": qc,          "higherBetter": True},
                {"id": "bc",    "label": "Big Chances",    "vals": bc_v,        "higherBetter": True},
            ],
        },
        "boxres": {
            "denying": [
                {"id": "opp_bea_cp",    "label": "Opp. Box Entries Attempted (Cross&Pass)",  "vals": opp_bea_cp,    "higherBetter": False, "context": "opp"},
                {"id": "opp_bea_carry", "label": "Opp. Box Entries Attempted (Carry)",       "vals": opp_bea_carry, "higherBetter": False, "context": "opp"},
                {"id": "opp_sbe",       "label": "Opp. Successful Box Entries",              "vals": opp_sbe,       "higherBetter": False, "context": "opp"},
                {"id": "opp_tib",       "label": "Opp. Touches in Box",                      "vals": opp_tib,       "higherBetter": False, "context": "opp"},
            ],
            "nullifying": [
                {"id": "opp_xg",    "label": "Opp. Total xG",         "vals": r(opp_xg, 3),    "higherBetter": False, "dec": 3, "context": "opp"},
                {"id": "opp_spxg",  "label": "Opp. Set Piece xG",     "vals": r(opp_spxg, 3),  "higherBetter": False, "dec": 3, "context": "opp"},
                {"id": "opp_shots", "label": "Opp. Shots",            "vals": opp_shots,       "higherBetter": False, "context": "opp"},
                {"id": "opp_xgps",  "label": "Opp. xG/Shot",          "vals": r(opp_xgps, 3),  "higherBetter": False, "dec": 3, "context": "opp"},
                {"id": "opp_psxg",  "label": "Opp. Post-Shot xG",     "vals": r(opp_psxg, 2),  "higherBetter": False, "dec": 2, "context": "opp"},
                {"id": "opp_qc",    "label": "Opp. Quality Chances",  "vals": opp_qc,          "higherBetter": False, "context": "opp"},
                {"id": "opp_bc",    "label": "Opp. Big Chances",      "vals": opp_bc_v,        "higherBetter": False, "context": "opp"},
            ],
        },
        "defint": {
            "volume": [
                {"id": "pressures",   "label": "Pressures (PA)",            "vals": r(pressures, 1),   "higherBetter": True, "dec": 1},
                {"id": "cp",          "label": "Counterpressures (PA)",     "vals": r(cp, 1),          "higherBetter": True, "dec": 1},
                {"id": "agg_actions", "label": "Aggressive Actions (PA)",   "vals": r(agg_actions, 1), "higherBetter": True, "dec": 1},
                {"id": "tack_int",    "label": "Tackles + Interceptions (PA)", "vals": r(tack_int, 1), "higherBetter": True, "dec": 1},
            ],
            "forward": [
                {"id": "press_oh",     "label": "Pressures in Opp. Half (PA)",          "vals": r(press_oh, 1),   "higherBetter": True, "dec": 1},
                {"id": "press_oh_pct", "label": "Pressures in Opp. Half %",             "vals": r(press_oh_pct, 4), "higherBetter": True, "pct": True, "dec": 3},
                {"id": "cp_oh",        "label": "Counterpressures in Opp. Half (PA)",   "vals": r(cp_oh, 1),      "higherBetter": True, "dec": 1},
            ],
            "progallowed": [
                {"id": "opp_dp",         "label": "Opp. Final 3rd Entries",        "vals": opp_dp,             "higherBetter": False, "context": "opp"},
                {"id": "opp_lb_comp_f3", "label": "Opp. LB Completed (Final Third)", "vals": opp_lb_comp_f3,   "higherBetter": False, "context": "opp"},
                {"id": "opp_lb_pct",     "label": "Opp. LB Completion % (F3)",     "vals": r(opp_lb_pct, 4),   "higherBetter": False, "pct": True, "dec": 3, "context": "opp"},
            ],
        },
        "obv": {
            "ncc": [
                {"id": "obv_total",   "label": "Total OBV",              "vals": r(obv_total, 3),   "higherBetter": True, "dec": 3},
                {"id": "obv_pass",    "label": "Pass OBV",               "vals": r(obv_pass, 3),    "higherBetter": True, "dec": 3},
                {"id": "obv_shot",    "label": "Shot OBV",               "vals": r(obv_shot, 3),    "higherBetter": True, "dec": 3},
                {"id": "obv_dribble", "label": "Dribble & Carry OBV",   "vals": r(obv_dribble, 3), "higherBetter": True, "dec": 3},
                {"id": "obv_def",     "label": "Defensive Action OBV",  "vals": r(obv_def, 3),     "higherBetter": True, "dec": 3},
            ],
            "opp": [
                {"id": "opp_obv_total",   "label": "Opp. Total OBV",             "vals": r(opp_obv_total, 3),   "higherBetter": False, "dec": 3, "context": "opp"},
                {"id": "opp_obv_pass",    "label": "Opp. Pass OBV",              "vals": r(opp_obv_pass, 3),    "higherBetter": False, "dec": 3, "context": "opp"},
                {"id": "opp_obv_shot",    "label": "Opp. Shot OBV",              "vals": r(opp_obv_shot, 3),    "higherBetter": False, "dec": 3, "context": "opp"},
                {"id": "opp_obv_dribble", "label": "Opp. Dribble & Carry OBV",  "vals": r(opp_obv_dribble, 3), "higherBetter": False, "dec": 3, "context": "opp"},
                {"id": "opp_obv_def",     "label": "Opp. Defensive Action OBV", "vals": r(opp_obv_def, 3),     "higherBetter": False, "dec": 3, "context": "opp"},
            ],
        },
    }

    info(f"Built DATA with {sum(len(g) for sec in DATA.values() for g in sec.values())} metric rows × {n_matches} matches")
    return DATA


BYPASS_SEGMENT_COLS = [
    "ByPassed Players 0 - 15 Min",
    "ByPassed Players 16 - 30 Min",
    "ByPassed Players 31 - 45 Min",
    "ByPassed Players 46 - 60 Min",
    "ByPassed Players 61 - 75 Min",
    "ByPassed Players 76 - 90 Min",
]
BYPASS_SEGMENT_LABELS = ["0–15'", "16–30'", "31–45'", "46–60'", "61–75'", "76–90'"]


def build_bypass_segments(matches, sheet_data):
    """Build a per-match × side × 6-window bypassed-players grid.

    Returns a list of dicts (one per match) with the shape:
        { "ncc": [v0_15, v16_30, ..., v76_90],
          "opp": [v0_15, v16_30, ..., v76_90] }
    Missing cells become None. Used by the new "Press Impact by 15-Min
    Window" section on the Packing & Progression subtab.
    """
    out = []
    for m in matches:
        gn = m["game_number"]
        entry = {"ncc": [], "opp": []}
        for side_key, side in (("ncc", "NCC"), ("opp", "OPP")):
            row = sheet_data.get((gn, side), {})
            for col in BYPASS_SEGMENT_COLS:
                v = row.get(col)
                entry[side_key].append(int(v) if v is not None else None)
        out.append(entry)
    return out


def render_bypass_segments_block(segments):
    """Render const BYPASS_SEGMENTS = [...]; for the dashboard.

    The BYPASS_SEGMENT_LABELS constant lives in the template and never
    changes per match, so we only swap the BYPASS_SEGMENTS array here.
    """
    lines = ["const BYPASS_SEGMENTS = ["]
    for e in segments:
        lines.append(
            "  { ncc:" + js_format(e["ncc"]) + ", opp:" + js_format(e["opp"]) + " },"
        )
    if lines[-1].endswith(","):
        lines[-1] = lines[-1].rstrip(",")
    lines.append("];")
    return "\n".join(lines)


def build_match_arrays(matches):
    """Produce MATCHES (compact) and MATCH_INFO (long) arrays for the dashboard."""
    SHORT_MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    matches_arr = []
    match_info_arr = []
    for m in matches:
        d = m["date"]
        short_date = f"{SHORT_MONTHS[d.month - 1]} {d.day}"
        long_date = f"{SHORT_MONTHS[d.month - 1]} {d.day}, {d.year}"
        label = f"GW{m['game_week']} \u00B7 M{m['game_number']}"   # · = U+00B7
        result = m["result"] or "—"
        # MATCH_INFO uses en-dash for results
        result_en = result.replace("-", "\u2013") if result != "—" else result
        matches_arr.append({
            "label": label, "date": short_date, "opp": m["opp"],
            "loc": m["loc"], "result": result,
        })
        match_info_arr.append({
            "label": label, "date": long_date, "opp": m["opp"],
            "loc": m["loc"], "result": result_en,
        })
    return matches_arr, match_info_arr


# ─────────────────────────────────────────────────────────────────────────────
# 3b) BUILD MATCH REPORT DATA (for Match Insights cards)
# ─────────────────────────────────────────────────────────────────────────────

def build_match_report_data(matches, sheet_data):
    """Build the MATCH_REPORT_DATA array consumed by the Match Insights tab.

    Most fields are read from the spreadsheet. Three fields are sourced from
    MANUAL_OVERRIDES (because they're not in the spreadsheet):
      ncc_sot / opp_sot          — Shots on Target
      ncc_press_regains / opp_press_regains — Press Regains
      top_shooters / top_kp / note — narrative strings
    """
    print(f"\n[3/3] Building MATCH_REPORT_DATA block")

    def sheet_val(gn, side, col):
        row = sheet_data.get((gn, side), {})
        return row.get(col)

    entries = []
    for m in matches:
        gn = m["game_number"]
        key = f"M{gn}"
        ov  = MANUAL_OVERRIDES.get(key, {})

        # Goals — from spreadsheet
        ncc_goals = sheet_val(gn, "NCC", "Goals") or 0
        opp_goals = sheet_val(gn, "OPP", "Goals") or 0

        # Result — derived from Goals in spreadsheet (computed in read_spreadsheet)
        match_result = m["result"]  # e.g. "W 2-1" or None
        if match_result:
            result_letter = match_result[0]
        elif ncc_goals > opp_goals:
            result_letter = "W"
        elif ncc_goals < opp_goals:
            result_letter = "L"
        else:
            result_letter = "D"

        # xG
        ncc_xg  = safe_round(sheet_val(gn, "NCC", "xG"), 3)
        opp_xg  = safe_round(sheet_val(gn, "OPP", "xG"), 3)
        ncc_xg_sp = safe_round(sheet_val(gn, "NCC", "Set Piece xG"), 3)
        opp_xg_sp = safe_round(sheet_val(gn, "OPP", "Set Piece xG"), 3)

        # Shots
        ncc_shots = sheet_val(gn, "NCC", "Shots")
        opp_shots = sheet_val(gn, "OPP", "Shots")

        # Shots on Target — manual only (not in spreadsheet)
        ncc_sot = ov.get("ncc_sot")
        opp_sot = ov.get("opp_sot")

        # Possession % — spreadsheet stores as integer (e.g. 62)
        ncc_poss = sheet_val(gn, "NCC", "Possession %")
        opp_poss = sheet_val(gn, "OPP", "Possession %")

        # Pass % — spreadsheet stores as decimal (e.g. 0.8051); convert to int %
        ncc_pass_raw = sheet_val(gn, "NCC", "Passing%")
        opp_pass_raw = sheet_val(gn, "OPP", "Passing%")
        ncc_pass_pct = round(float(ncc_pass_raw) * 100) if ncc_pass_raw is not None else None
        opp_pass_pct = round(float(opp_pass_raw) * 100) if opp_pass_raw is not None else None

        # Pressures
        ncc_pressures = sheet_val(gn, "NCC", "Pressures")
        opp_pressures = sheet_val(gn, "OPP", "Pressures")

        # Press Regains — manual only (not in spreadsheet)
        ncc_press_regains = ov.get("ncc_press_regains")
        opp_press_regains = ov.get("opp_press_regains")

        # Narrative fields — manual only
        top_shooters = ov.get("top_shooters")
        top_kp       = ov.get("top_kp")
        note         = ov.get("note")

        entry = {
            "ncc_goals": ncc_goals,
            "opp_goals": opp_goals,
            "result":    result_letter,
            "ncc_xg":   ncc_xg,
            "opp_xg":   opp_xg,
            "ncc_xg_sp": ncc_xg_sp,
            "opp_xg_sp": opp_xg_sp,
            "ncc_shots": ncc_shots,
            "opp_shots": opp_shots,
            "ncc_sot":   ncc_sot,
            "opp_sot":   opp_sot,
            "ncc_poss":  ncc_poss,
            "opp_poss":  opp_poss,
            "ncc_pass_pct": ncc_pass_pct,
            "opp_pass_pct": opp_pass_pct,
            "ncc_pressures": ncc_pressures,
            "opp_pressures": opp_pressures,
            "ncc_press_regains": ncc_press_regains,
            "opp_press_regains": opp_press_regains,
            "top_shooters": top_shooters,
            "top_kp":       top_kp,
            "note":         note,
        }
        entries.append(entry)
        info(f"M{gn} insights: result={result_letter} goals={ncc_goals}-{opp_goals} "
             f"xG={ncc_xg}-{opp_xg} sot={ncc_sot}-{opp_sot} poss={ncc_poss}%")

    return entries


def render_match_report_data_block(entries):
    """Render const MATCH_REPORT_DATA = [...] as JS."""
    lines = ["const MATCH_REPORT_DATA = ["]
    for e in entries:
        parts = []
        for k, v in e.items():
            parts.append(f"{k}:{js_format(v)}")
        lines.append("  { " + ", ".join(parts) + " },")
    if lines[-1].endswith(","):
        lines[-1] = lines[-1].rstrip(",")
    lines.append("];")
    return "\n".join(lines)


# ─────────────────────────────────────────────────────────────────────────────
# 4) RENDER HTML
# ─────────────────────────────────────────────────────────────────────────────

def js_format(value, indent=0):
    """Render a Python value as JavaScript (None → null, dicts → object literals)."""
    pad = "  " * indent
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        if isinstance(value, float):
            # Trim trailing zeros after decimal; but keep at least one
            s = repr(value)
            return s
        return str(value)
    if isinstance(value, str):
        # Escape backslashes and single quotes for JS single-quoted string
        s = value.replace("\\", "\\\\").replace("'", "\\'")
        return f"'{s}'"
    if isinstance(value, list):
        items = [js_format(v, indent) for v in value]
        return "[" + ",".join(items) + "]"
    if isinstance(value, dict):
        parts = [f"{k}:{js_format(v, indent)}" for k, v in value.items()]
        return "{" + ", ".join(parts) + "}"
    raise TypeError(f"Unsupported type: {type(value)}")


def render_matches_block(matches_arr):
    """Render the const MATCHES = [...] block."""
    lines = ["const MATCHES = ["]
    for m in matches_arr:
        lines.append(
            f"  {{ label: '{m['label']}', date: '{m['date']}', "
            f"opp: '{m['opp']}', loc: '{m['loc']}', result: '{m['result']}' }},"
        )
    # Remove trailing comma on last entry for cleanliness
    if lines[-1].endswith(","):
        lines[-1] = lines[-1].rstrip(",")
    lines.append("];")
    return "\n".join(lines)


def render_match_info_block(match_info_arr):
    """Render const MATCH_INFO."""
    lines = ["const MATCH_INFO = ["]
    for m in match_info_arr:
        lines.append(
            f"  {{ label: '{m['label']}', date: '{m['date']}', "
            f"opp: '{m['opp']}', loc: '{m['loc']}', result: '{m['result']}' }},"
        )
    if lines[-1].endswith(","):
        lines[-1] = lines[-1].rstrip(",")
    lines.append("];")
    return "\n".join(lines)


def render_data_block(DATA):
    """Render const DATA = {...} as readable JS."""
    lines = ["const DATA = {"]
    for section_key, section in DATA.items():
        lines.append(f"  {section_key}: {{")
        for group_key, rows in section.items():
            lines.append(f"    {group_key}: [")
            for row in rows:
                # Build property string in stable order matching original dashboard
                key_order = ["id", "label", "vals", "higherBetter", "pct", "dec", "context"]
                parts = []
                for k in key_order:
                    if k in row:
                        parts.append(f"{k}:{js_format(row[k])}")
                # Any extra keys not in standard order
                for k in row:
                    if k not in key_order:
                        parts.append(f"{k}:{js_format(row[k])}")
                lines.append("      { " + ", ".join(parts) + " },")
            # Trim trailing comma
            if lines[-1].endswith(","):
                lines[-1] = lines[-1].rstrip(",")
            lines.append("    ],")
        # Trim section trailing comma
        if lines[-1].endswith(","):
            lines[-1] = lines[-1].rstrip(",")
        lines.append("  },")
    if lines[-1].endswith(","):
        lines[-1] = lines[-1].rstrip(",")
    lines.append("};")
    return "\n".join(lines)


def replace_block(text, start_pattern, end_marker, new_content):
    """Replace from a line matching start_pattern through the next end_marker line."""
    lines = text.split("\n")
    start_idx = None
    for i, line in enumerate(lines):
        if start_pattern in line:
            start_idx = i
            break
    if start_idx is None:
        sys.exit(f"ERROR: Could not find start pattern: {start_pattern}")
    end_idx = None
    for i in range(start_idx, len(lines)):
        if end_marker in lines[i]:
            end_idx = i
            break
    if end_idx is None:
        sys.exit(f"ERROR: Could not find end marker '{end_marker}' after start at line {start_idx}")
    new_lines = lines[:start_idx] + new_content.split("\n") + lines[end_idx + 1:]
    return "\n".join(new_lines)


def render_html(matches_arr, match_info_arr, DATA, match_report_data, bypass_segments):
    print(f"\n  Rendering HTML")
    if not TEMPLATE.exists():
        sys.exit(f"ERROR: Template not found at {TEMPLATE}")
    html = TEMPLATE.read_text(encoding="utf-8")

    # Replace MATCHES block (lines that begin with "const MATCHES = [" through "];")
    new_matches = render_matches_block(matches_arr)
    html = replace_block(html, "const MATCHES = [", "];", new_matches)

    # Replace DATA block
    new_data = render_data_block(DATA)
    html = replace_block(html, "const DATA = {", "};", new_data)

    # Replace MATCH_INFO block
    new_match_info = render_match_info_block(match_info_arr)
    html = replace_block(html, "const MATCH_INFO = [", "];", new_match_info)

    # Replace MATCH_REPORT_DATA block (drives Match Insights cards)
    new_mrd = render_match_report_data_block(match_report_data)
    html = replace_block(html, "const MATCH_REPORT_DATA = [", "];", new_mrd)

    # Replace BYPASS_SEGMENTS block (drives Packing Momentum section).
    # The matching closing "];" is on a line by itself, not on the start line,
    # so this won't collide with the BYPASS_SEGMENT_LABELS one-liner above it.
    new_seg = render_bypass_segments_block(bypass_segments)
    html = replace_block(html, "const BYPASS_SEGMENTS = [", "];", new_seg)

    OUTPUT.write_text(html, encoding="utf-8")
    info(f"Wrote {OUTPUT.name} ({len(html):,} chars)")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("=" * 70)
    print("Courage Team Performance Dashboard — Build")
    print("=" * 70)
    matches, sheet_data = read_spreadsheet()
    DATA = build_data_block(matches, sheet_data)
    matches_arr, match_info_arr = build_match_arrays(matches)
    match_report_data = build_match_report_data(matches, sheet_data)
    bypass_segments = build_bypass_segments(matches, sheet_data)
    render_html(matches_arr, match_info_arr, DATA, match_report_data, bypass_segments)
    print("\n✅ Build complete.\n")
    print(f"   Output: {OUTPUT}")
    print(f"   Open this file in your browser to view the dashboard.\n")


if __name__ == "__main__":
    main()
